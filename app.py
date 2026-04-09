import os
import re
import socket
import subprocess
import threading
from datetime import datetime
from pathlib import Path

import qrcode
import openpyxl
from fastapi import FastAPI, Form, Request
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware


def get_local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "localhost"


def start_tunnel(port: int) -> str | None:
    """Open a localhost.run SSH tunnel. No account needed."""
    result = {"url": None}
    ready = threading.Event()

    def _run():
        try:
            proc = subprocess.Popen(
                ["ssh", "-o", "StrictHostKeyChecking=no",
                 "-R", f"80:localhost:{port}", "localhost.run"],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
            )
            for line in proc.stdout:
                match = re.search(r"https://[^\s]+\.lhr\.life", line)
                if match:
                    result["url"] = match.group(0)
                    ready.set()
        except Exception as e:
            print(f"[tunnel] {e}")
            ready.set()

    threading.Thread(target=_run, daemon=True).start()
    ready.wait(timeout=15)
    return result["url"]


_port = int(os.getenv("PORT", 8000))
_explicit_url = os.getenv("BASE_URL")

if _explicit_url:
    BASE_URL = _explicit_url
else:
    print("[tunnel] Starting public tunnel via localhost.run...")
    _tunnel_url = start_tunnel(_port)
    if _tunnel_url:
        print(f"[tunnel] Public URL: {_tunnel_url}")
        BASE_URL = _tunnel_url
    else:
        print("[tunnel] Falling back to local IP")
        BASE_URL = f"http://{get_local_ip()}:{_port}"

APP_DIR = Path(__file__).parent
EXCEL_PATH = APP_DIR / "responses.xlsx"
QR_PATH = APP_DIR / "static" / "qr.png"
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
SECRET_KEY = os.getenv("SECRET_KEY", "change-me-in-production")

QUESTIONS = [
    "What is your name?",
    "How would you describe your overall experience?",
    "What did you like most?",
    "What could be improved?",
    "Any additional comments or suggestions?",
]

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)
app.mount("/static", StaticFiles(directory=str(APP_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))


def is_admin(request: Request) -> bool:
    return request.session.get("is_admin") is True


def generate_qr():
    img = qrcode.make(f"{BASE_URL}/survey")
    img.save(QR_PATH)


def save_response(answers: list[str]):
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"
        headers = ["Timestamp"] + [f"Q{i+1}: {q}" for i, q in enumerate(QUESTIONS)]
        ws.append(headers)
        for col in ws.iter_cols(1, len(headers), 1, 1):
            for cell in col:
                cell.font = openpyxl.styles.Font(bold=True)

    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S")] + answers)
    wb.save(EXCEL_PATH)


@app.on_event("startup")
async def startup():
    generate_qr()


@app.get("/")
async def root():
    return RedirectResponse(url="/survey")


@app.get("/survey")
async def survey(request: Request):
    return templates.TemplateResponse(
        request=request, name="survey.html", context={"questions": QUESTIONS}
    )


@app.post("/submit")
async def submit(
    request: Request,
    q1: str = Form(""),
    q2: str = Form(""),
    q3: str = Form(""),
    q4: str = Form(""),
    q5: str = Form(""),
):
    save_response([q1, q2, q3, q4, q5])
    return RedirectResponse(url="/thankyou", status_code=303)


@app.get("/thankyou")
async def thankyou(request: Request):
    return templates.TemplateResponse(request=request, name="thankyou.html")


# ── Admin ────────────────────────────────────────────────────────────────────

@app.get("/admin")
async def admin(request: Request):
    count = 0
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        count = wb.active.max_row - 1
    return templates.TemplateResponse(
        request=request,
        name="admin.html",
        context={
            "survey_url": f"{BASE_URL}/survey",
            "response_count": count,
            "is_public": BASE_URL.startswith("https://"),
            "download_error": None,
        },
    )


@app.post("/download")
async def download(request: Request, password: str = Form("")):
    if password != ADMIN_PASSWORD:
        count = 0
        if EXCEL_PATH.exists():
            wb = openpyxl.load_workbook(EXCEL_PATH)
            count = wb.active.max_row - 1
        return templates.TemplateResponse(
            request=request,
            name="admin.html",
            context={
                "survey_url": f"{BASE_URL}/survey",
                "response_count": count,
                "is_public": BASE_URL.startswith("https://"),
                "download_error": "Incorrect password.",
            },
        )
    if not EXCEL_PATH.exists():
        return {"error": "No responses yet."}
    return FileResponse(
        path=str(EXCEL_PATH),
        filename="survey_responses.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
